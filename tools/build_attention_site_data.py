import csv
import json
import shutil
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_PANEL_ROOT = ROOT / "all_patient_attention_results_morphology_guided_filtered"
SOURCE_CASE_ROOT = ROOT / "case_csvs_all"
TARGET_IMAGE_ROOT = REPO_ROOT / "img" / "cases"
TARGET_DATA_DIR = REPO_ROOT / "data"
CLINICAL_WORKBOOK = ROOT / "os_based_merge_pMMR.clean.matched.xlsx"

RISK_GROUP_IDS = ("high_risk", "low_risk")


def export_webp(src: Path, dst_without_suffix: Path, quality: int) -> Path:
    dst_without_suffix.parent.mkdir(parents=True, exist_ok=True)
    dst = dst_without_suffix.with_suffix(".webp")
    with Image.open(src) as img:
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGBA")
        img.save(dst, format="WEBP", quality=quality, method=6)
    return dst


def read_json(path: Path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_top_scores(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    def top_row(key: str):
        valid = [row for row in rows if row.get("x") != "-1" and row.get("y") != "-1"]
        best = max(valid, key=lambda row: float(row[key]))
        return {
            "x": int(float(best["x"])),
            "y": int(float(best["y"])),
            "percentile": round(float(best[key]), 4),
        }

    return {
        "pathology_self_percentile": top_row("pathology_self_percentile"),
        "clinical_guided_percentile": top_row("clinical_guided_percentile"),
        "cell_guided_percentile": top_row("cell_guided_percentile"),
    }


def read_clinical_data(path: Path):
    if not path.exists():
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    clinical = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        patient_id = str(row[0]).strip()
        time_value = row[1]
        t_value = row[3]
        n_value = row[4]
        m_value = row[5]
        age_value = row[6]
        gleason_pattern = row[7]
        gleason_score = row[8]

        def with_prefix(prefix: str, value):
            if value in (None, ""):
                return None
            value_str = str(value).strip()
            return f"{prefix}{value_str}"

        def format_age(value):
            if value in (None, ""):
                return None
            numeric = float(value)
            return int(numeric) if numeric.is_integer() else round(numeric, 1)

        def format_survival(value):
            if value in (None, ""):
                return None
            numeric = float(value)
            return int(numeric) if numeric.is_integer() else round(numeric, 1)

        def format_gleason(pattern, score):
            if pattern not in (None, "") and score not in (None, ""):
                score_num = float(score)
                score_text = int(score_num) if score_num.is_integer() else round(score_num, 1)
                return f"{str(pattern).strip()} ({score_text})"
            if pattern not in (None, ""):
                return str(pattern).strip()
            if score not in (None, ""):
                score_num = float(score)
                return int(score_num) if score_num.is_integer() else round(score_num, 1)
            return None

        clinical[patient_id] = {
            "age": format_age(age_value),
            "tStage": with_prefix("T", t_value),
            "nStage": with_prefix("N", n_value),
            "mStage": with_prefix("M", m_value),
            "gleason": format_gleason(gleason_pattern, gleason_score),
            "survival": format_survival(time_value),
        }

    return clinical


def discover_case_groups():
    case_groups = {}
    for group in RISK_GROUP_IDS:
        group_dir = SOURCE_PANEL_ROOT / group
        case_ids = sorted([item.name for item in group_dir.iterdir() if item.is_dir()])
        case_groups[group] = case_ids
    return case_groups


def build_case(group: str, patient_id: str, clinical_lookup):
    panel_dir = SOURCE_PANEL_ROOT / group / patient_id
    case_dir = SOURCE_CASE_ROOT / patient_id

    meta_path = case_dir / f"{patient_id}_guided_attention_meta.json"
    csv_path = case_dir / f"{patient_id}_guided_patch_attention_scores.csv"

    image_sources = {
        "panel": (panel_dir / f"{patient_id}_interpretability_panel.png", 76),
        "wholeSlide": (panel_dir / f"{patient_id}_whole_slide_crop.png", 74),
        "clinicalHeatmap": (panel_dir / f"{patient_id}_clinical_guided_heatmap.png", 70),
        "cellHeatmap": (panel_dir / f"{patient_id}_cell_guided_heatmap.png", 70),
        "clinicalPercentile": (case_dir / f"{patient_id}_clinical_guided_percentile.png", 68),
        "cellPercentile": (case_dir / f"{patient_id}_cell_guided_percentile.png", 68),
        "pathologySelf": (case_dir / f"{patient_id}_pathology_self_percentile.png", 68),
    }

    target_dir = TARGET_IMAGE_ROOT / patient_id
    image_paths = {}
    for key, (src, quality) in image_sources.items():
        if src.exists():
            dst = export_webp(src, target_dir / src.stem, quality)
            image_paths[key] = f"img/cases/{patient_id}/{dst.name}"

    meta = read_json(meta_path)
    top_scores = read_top_scores(csv_path)
    clinical_info = clinical_lookup.get(patient_id, {})
    cell_groups = [group_name for group_name in meta["cell_group_names"] if group_name != "other"]

    return {
        "id": patient_id,
        "label": patient_id,
        "riskGroup": group,
        "patientId": patient_id,
        "clinicalInfo": {
            "age": clinical_info.get("age"),
            "tStage": clinical_info.get("tStage"),
            "nStage": clinical_info.get("nStage"),
            "mStage": clinical_info.get("mStage"),
            "gleason": clinical_info.get("gleason"),
            "survival": clinical_info.get("survival"),
        },
        "fold": f"fold{meta['fold_id']}",
        "patchCount": meta["patch_count"],
        "patchSize": f"{meta['stride']} x {meta['stride']}",
        "featureDim": meta["feature_dim"],
        "gridShape": f"{meta['grid_shape'][0]} x {meta['grid_shape'][1]}",
        "clinicalTokenCount": len(meta["clinical_cols"]),
        "cellGroupCount": len(cell_groups),
        "cellGroups": cell_groups,
        "images": image_paths,
        "topScores": top_scores,
    }


def main():
    TARGET_DATA_DIR.mkdir(parents=True, exist_ok=True)
    if TARGET_IMAGE_ROOT.exists():
        shutil.rmtree(TARGET_IMAGE_ROOT)

    case_groups = discover_case_groups()
    clinical_lookup = read_clinical_data(CLINICAL_WORKBOOK)
    cases = []
    for group, patient_ids in case_groups.items():
        for patient_id in patient_ids:
            cases.append(build_case(group, patient_id, clinical_lookup))

    payload = {
        "paperTitle": "BiCrossSurv attention visualization for prostate cancer prognosis",
        "paperDescription": "MedAILab project site featuring BiCrossSurv cross-modal attention results for prostate cancer prognosis, with access to related multimodal pathology demos and archived project pages.",
        "riskGroups": [
            {"id": "all", "label": "All cases"},
            {"id": "high_risk", "label": "High risk"},
            {"id": "low_risk", "label": "Low risk"},
        ],
        "cases": cases,
    }

    with (TARGET_DATA_DIR / "cases.json").open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
